r"""
资产负债 Excel 解析模块

数据源优先级:
  1. 环境变量 JINKUI_EXCEL_PATH
  2. config.json 中的 excel_path 字段
  3. 项目根目录下的 资产负债.xlsx
结构: 每年一个 sheet，包含资产、负债、净资产及分析区域
"""
import json
import os
import re
import math
import logging
from datetime import datetime, timedelta
from typing import Any

import openpyxl

_PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))


def _resolve_excel_path() -> str:
    """解析 Excel 文件路径，按优先级查找"""
    # 1. 环境变量
    env_path = os.environ.get("JINKUI_EXCEL_PATH")
    if env_path and os.path.isfile(env_path):
        return env_path

    # 2. config.json
    config_file = os.path.join(_PROJECT_DIR, "config.json")
    if os.path.isfile(config_file):
        try:
            with open(config_file, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            cfg_path = cfg.get("excel_path", "")
            if cfg_path and os.path.isfile(cfg_path):
                return cfg_path
        except (json.JSONDecodeError, OSError):
            pass

    # 3. 项目目录下默认文件名
    default_name = os.path.join(_PROJECT_DIR, "资产负债.xlsx")
    if os.path.isfile(default_name):
        return default_name

    raise FileNotFoundError(
        "找不到资产负债表 Excel 文件。请:\n"
        "  1. 将文件放在项目目录并命名为 资产负债.xlsx, 或\n"
        "  2. 在 config.json 中设置 excel_path, 或\n"
        "  3. 设置环境变量 JINKUI_EXCEL_PATH"
    )


EXCEL_PATH = _resolve_excel_path()


def _excel_serial_to_date(serial: int | float) -> datetime | None:
    """将 Excel 日期序列号转为 datetime"""
    if serial is None or serial <= 0:
        return None
    base = datetime(1899, 12, 30)
    return base + timedelta(days=int(serial))


def _parse_month_header(value: Any) -> datetime | None:
    """解析月份表头，返回该月第一天的 datetime"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return _excel_serial_to_date(int(value))
    s = str(value).strip()
    m = re.match(r"(\d{4})年([一二三四五六七八九十]+)月?", s)
    if m:
        year = int(m.group(1))
        cn_map = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6,
                  "七": 7, "八": 8, "九": 9, "十": 10, "十一": 11, "十二": 12}
        month = cn_map.get(m.group(2))
        if month:
            return datetime(year, month, 1)
    m = re.match(r"(\d{4})[/年-](\d{1,2})", s)
    if m:
        return datetime(int(m.group(1)), int(m.group(2)), 1)
    return None


def _to_float(val: Any) -> float | None:
    """安全转为 float"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if math.isnan(val) or math.isinf(val):
            return None
        return float(val)
    if isinstance(val, str):
        try:
            return float(val)
        except ValueError:
            return None
    return None


def _clean_name(val: Any) -> str:
    """清洗账户名称"""
    if val is None:
        return ""
    return str(val).strip()


# 已知的汇总/标签行名称（用于排除）
_SKIP_NAMES = {"", "资产", "資產", "负债", "負債", "合计", "合計", "总计", "總計",
               "汇总", "匯總", "小计", "小計",
               "净资产", "淨資產", "净资产变化", "淨資產變化",
               "收入", "實際收入", "实际收入", "实际支出", "實際支出",
               "计划支出", "計劃支出", "计划储蓄", "計劃儲蓄", "收支差额", "收支差額"}


def _is_skip_name(name: str) -> bool:
    return name in _SKIP_NAMES or name in ("备注", "注释")


# --- 带文件修改时间检测的缓存 ---
_cached_data: dict | None = None
_cached_mtime: float | None = None


def load_data() -> dict:
    """加载完整数据（自动检测 Excel 文件修改，变化时重载）"""
    global _cached_data, _cached_mtime
    try:
        current_mtime = os.path.getmtime(EXCEL_PATH)
    except OSError:
        current_mtime = None

    if _cached_data is not None and current_mtime is not None and current_mtime <= _cached_mtime:
        return _cached_data

    _cached_data = _load_data_impl()
    _cached_mtime = current_mtime
    return _cached_data


def refresh_data():
    """强制刷新缓存（Excel 更新后调用）"""
    global _cached_data, _cached_mtime
    _cached_data = None
    _cached_mtime = None


def _load_data_impl() -> dict:
    """
    加载完整数据（实际实现）。
    返回结构见函数末尾 return。
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # 只处理纯数字名称的工作表（年份），跳过 "Sheet1"、"目录" 等
    years = []
    for s in wb.sheetnames:
        s_clean = s.strip()
        # 仅接受纯 ASCII 数字（排除 ¹, ① 等 Unicode digit）
        if s_clean.isascii() and s_clean.isdigit():
            years.append(s_clean)
        else:
            logging.warning("跳过非年份工作表: '%s'", s)
    if not years:
        raise ValueError(
            f"Excel 中没有可识别的年份工作表（找到: {wb.sheetnames}）"
        )
    years = sorted(years, key=int)

    all_months: list[str] = []            # "YYYY-MM"
    accounts_data: dict[str, list] = {}    # account_name -> [val or None ...]
    income_list: list[float | None] = []
    actual_expense_list: list[float | None] = []
    planned_expense_list: list[float | None] = []
    net_worth_change_list: list[float | None] = []
    monthly_detail: list[dict] = []

    for year_str in years:
        ws = wb[year_str]
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
        if not rows:
            continue

        # --- Parse month headers ---
        headers = rows[0]
        month_dates: list[datetime] = []
        for idx in range(1, len(headers)):
            val = headers[idx]
            # Stop at first None/empty header after we have at least one valid month
            if val is None and month_dates:
                break
            dt = _parse_month_header(val)
            if dt is None and month_dates:
                break  # Stop at unrecognized header (sheet may have extra garbage columns)
            if dt is None:
                continue  # Skip unrecognized before first valid month
            # Ensure no duplicate month within same year
            if month_dates and dt <= month_dates[-1]:
                continue  # Skip duplicates
            month_dates.append(dt)

        n = len(month_dates)  # number of month columns

        # --- Scan rows to classify each ---
        # Each row gets a type: 'asset', 'liability', 'separator', 'networth', 'analysis', 'ignore'
        section = "asset"  # current section
        asset_rows = []
        liability_rows = []
        networth_start = len(rows)  # row index where net worth section begins

        for i in range(1, len(rows)):
            row = rows[i]
            first = _clean_name(row[0])

            # Check if all cells are None
            all_none = all(c is None for c in row)

            if all_none:
                continue

            if first in ("负债", "負債"):
                section = "liability"
                continue
            if first in ("净资产", "淨資產"):
                section = "networth"
                networth_start = i
                continue
            if _is_skip_name(first):
                continue

            if section == "asset" and first:
                asset_rows.append((i, first, row))
            elif section == "liability" and first:
                liability_rows.append((i, first, row))

        # --- Extract asset values ---
        asset_vals_by_account: dict[str, list] = {}
        for _, name, row in asset_rows:
            vals = [_to_float(row[j]) for j in range(1, min(1 + n, len(row)))]
            vals = vals + [None] * (n - len(vals))
            # Deduplicate: if same name appears, use the one with more non-None values
            if name in asset_vals_by_account:
                old = asset_vals_by_account[name]
                new_non_none = sum(1 for v in vals if v is not None)
                old_non_none = sum(1 for v in old if v is not None)
                if new_non_none > old_non_none:
                    asset_vals_by_account[name] = vals
            else:
                asset_vals_by_account[name] = vals

        # --- Extract liability values ---
        liability_vals_by_account: dict[str, list] = {}
        for _, name, row in liability_rows:
            vals = [_to_float(row[j]) for j in range(1, min(1 + n, len(row)))]
            vals = vals + [None] * (n - len(vals))
            if name in liability_vals_by_account:
                old = liability_vals_by_account[name]
                new_non_none = sum(1 for v in vals if v is not None)
                old_non_none = sum(1 for v in old if v is not None)
                if new_non_none > old_non_none:
                    liability_vals_by_account[name] = vals
            else:
                liability_vals_by_account[name] = vals

        # --- Compute totals ourselves (don't rely on formula cache) ---
        total_assets_vals = []
        total_liabilities_vals = []
        for mi in range(n):
            ta = sum(v[mi] for v in asset_vals_by_account.values() if mi < len(v) and v[mi] is not None) or 0
            tl = sum(v[mi] for v in liability_vals_by_account.values() if mi < len(v) and v[mi] is not None) or 0
            total_assets_vals.append(ta)
            total_liabilities_vals.append(tl)

        # --- Extract analysis rows ---
        inc_vals = _find_analysis_row(rows, n, ["收入", "實際收入", "实际收入"], networth_start)
        exp_vals = _find_analysis_row(rows, n, ["实际支出", "實際支出"], networth_start)
        plan_vals = _find_analysis_row(rows, n, ["计划支出", "計劃支出"], networth_start)
        change_vals = _find_analysis_row(rows, n, ["净资产变化", "淨資產變化"], networth_start)

        # --- Build monthly records ---
        for mi in range(n):
            month_str = month_dates[mi].strftime("%Y-%m")
            all_months.append(month_str)

            nw = total_assets_vals[mi] - total_liabilities_vals[mi]

            month_assets = {k: v[mi] for k, v in asset_vals_by_account.items()
                           if mi < len(v) and v[mi] is not None}
            month_liabilities = {k: v[mi] for k, v in liability_vals_by_account.items()
                                if mi < len(v) and v[mi] is not None}

            monthly_detail.append({
                "date": month_str,
                "year": month_dates[mi].year,
                "month": month_dates[mi].month,
                "assets": month_assets,
                "liabilities": month_liabilities,
                "total_assets": round(total_assets_vals[mi], 2),
                "total_liabilities": round(total_liabilities_vals[mi], 2),
                "net_worth": round(nw, 2),
                "income": round(inc_vals[mi], 2) if mi < len(inc_vals) and inc_vals[mi] is not None else None,
                "actual_expense": round(exp_vals[mi], 2) if mi < len(exp_vals) and exp_vals[mi] is not None else None,
                "planned_expense": round(plan_vals[mi], 2) if mi < len(plan_vals) and plan_vals[mi] is not None else None,
                "net_worth_change": round(change_vals[mi], 2) if mi < len(change_vals) and change_vals[mi] is not None else None,
            })

        # --- Extend global account lists (detect name conflicts) ---
        def _merge_accounts(by_account: dict, section_label: str):
            conflict_names = set(asset_vals_by_account) & set(liability_vals_by_account)
            for name, vals in by_account.items():
                key = f"{name}（{section_label}）" if name in conflict_names else name
                if key not in accounts_data:
                    accounts_data[key] = [None] * (len(all_months) - n)
                accounts_data[key].extend(vals)
        _merge_accounts(asset_vals_by_account, "资产")
        _merge_accounts(liability_vals_by_account, "负债")
        # Pad accounts not in this year
        for name in accounts_data:
            if len(accounts_data[name]) < len(all_months):
                accounts_data[name].extend([None] * (len(all_months) - len(accounts_data[name])))

    wb.close()

    # --- Build summary globals ---
    total_assets_all = [m["total_assets"] for m in monthly_detail]
    total_liabilities_all = [m["total_liabilities"] for m in monthly_detail]
    net_worth_all = [m["net_worth"] for m in monthly_detail]
    income_all = [m["income"] for m in monthly_detail]
    actual_expense_all = [m["actual_expense"] for m in monthly_detail]
    planned_expense_all = [m["planned_expense"] for m in monthly_detail]
    net_worth_change_all = [m["net_worth_change"] for m in monthly_detail]

    # Latest non-None net worth index（跳过全零的占位月份）
    latest_idx = -1  # -1 表示无有效数据
    for i in range(len(net_worth_all) - 1, -1, -1):
        if net_worth_all[i] is not None and (total_assets_all[i] != 0 or total_liabilities_all[i] != 0):
            latest_idx = i
            break

    return {
        "years": years,
        "months": all_months,
        "month_labels": all_months,
        "accounts": accounts_data,
        "total_assets": total_assets_all,
        "total_liabilities": total_liabilities_all,
        "net_worth": net_worth_all,
        "income": income_all,
        "actual_expense": actual_expense_all,
        "planned_expense": planned_expense_all,
        "net_worth_change": net_worth_change_all,
        "monthly": monthly_detail,
        "latest_idx": latest_idx,
    }


def _find_analysis_row(rows: list, n_months: int, labels: list[str], start_from: int = 0) -> list:
    """在 rows 中查找匹配 label 的分析行（值可能在同行的后续列或下一行）。
    从 start_from 行开始搜索（用于跳过资产/负债区域）。
    """
    for i in range(start_from, len(rows)):
        row = rows[i]
        first = _clean_name(row[0])
        if first in labels:
            vals = [_to_float(row[j]) for j in range(1, min(1 + n_months, len(row)))]
            if any(v is not None for v in vals):
                return vals + [None] * (n_months - len(vals))
            # Try next row (the label row may have the label, values are in cells below)
            if i + 1 < len(rows):
                next_row = rows[i + 1]
                if not _clean_name(next_row[0]) or _clean_name(next_row[0]) == "":
                    vals2 = [_to_float(next_row[j]) for j in range(1, min(1 + n_months, len(next_row)))]
                    if any(v is not None for v in vals2):
                        return vals2 + [None] * (n_months - len(vals2))
            # Try two rows below (some sheets have label row + empty + values)
            if i + 2 < len(rows):
                row3 = rows[i + 2]
                if not _clean_name(row3[0]) or _clean_name(row3[0]) == "":
                    vals3 = [_to_float(row3[j]) for j in range(1, min(1 + n_months, len(row3)))]
                    if any(v is not None for v in vals3):
                        return vals3 + [None] * (n_months - len(vals3))
    return [None] * n_months


def get_summary() -> dict:
    """获取首页摘要——最新月份概览"""
    data = load_data()
    idx = data["latest_idx"]
    if idx < 0 or not data["monthly"]:
        return {"date": "", "total_assets": 0, "total_liabilities": 0, "net_worth": 0}
    monthly = data["monthly"]
    latest = monthly[idx] if idx < len(monthly) else {}
    prev = monthly[idx - 1] if idx > 0 else None

    return {
        "date": latest.get("date", ""),
        "total_assets": latest.get("total_assets", 0),
        "total_liabilities": latest.get("total_liabilities", 0),
        "net_worth": latest.get("net_worth", 0),
        "income": latest.get("income"),
        "actual_expense": latest.get("actual_expense"),
        "planned_expense": latest.get("planned_expense"),
        "net_worth_change": latest.get("net_worth_change"),
        "assets_breakdown": latest.get("assets", {}),
        "liabilities_breakdown": latest.get("liabilities", {}),
        "prev_net_worth": prev.get("net_worth") if prev else None,
    }


def get_trend(years: list[str] | None = None) -> dict:
    """获取净资产及资产/负债趋势"""
    data = load_data()
    months, nw, ta, tl = [], [], [], []
    for i, m in enumerate(data["months"]):
        y = m[:4]
        if years is None or y in years:
            months.append(data["month_labels"][i])
            nw.append(data["net_worth"][i])
            ta.append(data["total_assets"][i])
            tl.append(data["total_liabilities"][i])
    return {
        "months": months,
        "net_worth": nw,
        "total_assets": ta,
        "total_liabilities": tl,
    }


def get_asset_breakdown(latest_only: bool = True) -> dict:
    """获取资产/负债构成"""
    data = load_data()
    if latest_only:
        idx = data["latest_idx"]
        if idx < 0:
            return {"assets": {}, "liabilities": {}}
        latest = data["monthly"][idx] if idx < len(data["monthly"]) else {}
        return {
            "date": latest.get("date"),
            "assets": latest.get("assets", {}),
            "liabilities": latest.get("liabilities", {}),
        }
    return {
        "months": data["month_labels"],
        "assets_by_month": [{k: v for k, v in m["assets"].items()} for m in data["monthly"]],
        "liabilities_by_month": [{k: v for k, v in m["liabilities"].items()} for m in data["monthly"]],
    }


def get_monthly_detail(year: int, month: int) -> dict | None:
    """获取指定年月的完整明细"""
    data = load_data()
    for m in data["monthly"]:
        if m["year"] == year and m["month"] == month:
            return m
    return None


def get_available_years() -> list[str]:
    """获取可用年份列表（复用缓存）"""
    data = load_data()
    return data["years"]


# ═══════════════════════════════════════════
#  报表功能
# ═══════════════════════════════════════════

def get_monthly_report(year: int, month: int) -> dict | None:
    """生成指定月份的财务报告"""
    detail = get_monthly_detail(year, month)
    if detail is None:
        return None

    prev = None
    if detail["month"] > 1:
        prev = get_monthly_detail(year, detail["month"] - 1)
    elif year > int(get_available_years()[0]):
        prev = get_monthly_detail(year - 1, 12)

    # 同比: 去年同月
    last_year = get_monthly_detail(year - 1, detail["month"]) if year - 1 >= int(get_available_years()[0]) else None

    # 资产/负债排序
    assets_sorted = sorted(detail["assets"].items(), key=lambda x: -x[1])
    liabilities_sorted = sorted(detail["liabilities"].items(), key=lambda x: -x[1])

    income = detail.get("income")
    expense = detail.get("actual_expense")
    planned = detail.get("planned_expense")
    savings_rate = None
    if income is not None and expense is not None and income > 0:
        savings_rate = round((income - expense) / income * 100, 1)

    budget_diff = None
    budget_pct = None
    if expense is not None and planned is not None and planned > 0:
        budget_diff = round(expense - planned, 2)
        budget_pct = round(expense / planned * 100, 1)

    return {
        "date": detail["date"],
        "year": detail["year"],
        "month": detail["month"],
        "income": income,
        "expense": expense,
        "planned_expense": planned,
        "savings_rate": savings_rate,
        "net_worth": detail["net_worth"],
        "total_assets": detail["total_assets"],
        "total_liabilities": detail["total_liabilities"],
        "net_worth_change": detail.get("net_worth_change"),
        "prev_net_worth": prev["net_worth"] if prev else None,
        "yoy_net_worth": last_year["net_worth"] if last_year else None,
        "budget_diff": budget_diff,
        "budget_pct": budget_pct,
        "top_assets": assets_sorted[:8],
        "top_liabilities": liabilities_sorted[:8],
    }


def get_annual_report(year: int) -> dict | None:
    """生成年度财务报告"""
    data = load_data()
    year_str = str(year)
    months_in_year = [m for m in data["monthly"] if str(m["year"]) == year_str]
    if not months_in_year:
        return None

    first, last = months_in_year[0], months_in_year[-1]

    total_income = sum(m.get("income") or 0 for m in months_in_year)
    total_expense = sum(m.get("actual_expense") or 0 for m in months_in_year)
    savings_total = total_income - total_expense
    avg_savings_rate = round(savings_total / total_income * 100, 1) if total_income > 0 else 0

    # 年初年末净资产
    nw_start = first["net_worth"]
    nw_end = last["net_worth"]
    nw_growth_pct = round((nw_end - nw_start) / nw_start * 100, 1) if nw_start > 0 else 0

    # 找净资产最大/最小月（显式 None 检查，避免 0 被视为 falsy）
    best_month = max(months_in_year, key=lambda m: m.get("net_worth_change") if m.get("net_worth_change") is not None else -float("inf"))
    worst_month = min(months_in_year, key=lambda m: m.get("net_worth_change") if m.get("net_worth_change") is not None else float("inf"))

    # 预算执行（显式 None 检查，0 支出是合法值）
    budget_ok = sum(1 for m in months_in_year
                    if m.get("actual_expense") is not None and m.get("planned_expense") is not None
                    and m["actual_expense"] <= m["planned_expense"])
    budget_total = sum(1 for m in months_in_year if m.get("planned_expense") is not None)

    # 同比（从 yoy_data 查找去年末净资产，避免递归重建全量报告）
    yoy_list = get_yoy_data()
    prev_nw_end = None
    for entry in yoy_list:
        if entry["year"] == year - 1:
            prev_nw_end = entry["nw_end"]
            break

    return {
        "year": year,
        "months_count": len(months_in_year),
        "total_income": round(total_income, 2),
        "total_expense": round(total_expense, 2),
        "savings_total": round(savings_total, 2),
        "avg_savings_rate": avg_savings_rate,
        "nw_start": nw_start,
        "nw_end": nw_end,
        "nw_growth": round(nw_end - nw_start, 2),
        "nw_growth_pct": nw_growth_pct,
        "best_month": {"date": best_month["date"], "change": best_month.get("net_worth_change")},
        "worst_month": {"date": worst_month["date"], "change": worst_month.get("net_worth_change")},
        "budget_compliance": f"{budget_ok}/{budget_total}" if budget_total > 0 else "N/A",
        "yoy_nw": prev_nw_end,
        "yoy_growth_pct": round((nw_end - prev_nw_end) / prev_nw_end * 100, 1)
            if prev_nw_end is not None and prev_nw_end > 0 else None,
    }


def get_yoy_data() -> list[dict]:
    """获取多年同比数据（每年末净资产）"""
    data = load_data()
    result = []
    for year_str in data["years"]:
        months_in_year = [m for m in data["monthly"] if str(m["year"]) == year_str]
        if not months_in_year:
            continue
        last = months_in_year[-1]
        # 找全年的收入合计
        total_income = sum(m.get("income") or 0 for m in months_in_year)
        total_expense = sum(m.get("actual_expense") or 0 for m in months_in_year)
        first = months_in_year[0]
        result.append({
            "year": int(year_str),
            "nw_end": last["net_worth"],
            "nw_start": first["net_worth"],
            "total_assets": last["total_assets"],
            "total_liabilities": last["total_liabilities"],
            "total_income": round(total_income, 2),
            "total_expense": round(total_expense, 2),
        })
    return result


if __name__ == "__main__":
    d = load_data()
    print(f"Years: {d['years']}")
    print(f"Total months: {len(d['months'])}")
    print(f"Latest index: {d['latest_idx']}")
    print(f"Latest month: {d['monthly'][d['latest_idx']]['date']}")
    s = get_summary()
    print(f"\n=== Latest Summary ===")
    print(f"  Total Assets:  {s['total_assets']:>12,.2f}")
    print(f"  Total Liabilities: {s['total_liabilities']:>8,.2f}")
    print(f"  Net Worth:     {s['net_worth']:>12,.2f}")
    print(f"  Income:        {s['income']}")
    print(f"  Actual Expense:{s['actual_expense']}")
    print(f"  Planned Expense:{s['planned_expense']}")
    print(f"\n  Assets breakdown ({len(s['assets_breakdown'])} accounts):")
    for k, v in sorted(s['assets_breakdown'].items(), key=lambda x: -x[1] if x[1] else 0):
        print(f"    {k}: {v:,.2f}")
    print(f"\n  Liabilities breakdown ({len(s['liabilities_breakdown'])} accounts):")
    for k, v in s['liabilities_breakdown'].items():
        print(f"    {k}: {v:,.2f}")
    print(f"\n  Accounts tracked: {list(d['accounts'].keys())}")
